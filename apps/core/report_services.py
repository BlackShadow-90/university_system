"""
PDF and Report Generation Services
"""
import io
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP

from django.http import HttpResponse
from django.utils.translation import gettext as _
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    Image, PageBreak, KeepTogether
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from apps.enrollments.models import Enrollment
from apps.students.models import Student
from apps.courses.models import Course

# Import unified calculation services
from apps.core.academic_services import (
    FinalScoreCalculationService,
    GPACalculationService,
    AcademicRecalculationService,
)


class PDFGenerator:
    """Base PDF generation utilities"""
    
    @staticmethod
    def create_response(filename):
        """Create HTTP response with PDF content type"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @staticmethod
    def create_styles():
        """Create custom paragraph styles for documents"""
        styles = getSampleStyleSheet()
        
        # Title style
        styles.add(ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))
        
        # Heading 2
        styles.add(ParagraphStyle(
            'CustomHeading2',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        ))
        
        # Normal text
        styles.add(ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=6,
            fontName='Helvetica'
        ))
        
        # Small text
        styles.add(ParagraphStyle(
            'Small',
            parent=styles['Normal'],
            fontSize=8,
            spaceAfter=4,
            fontName='Helvetica'
        ))
        
        # Right aligned
        styles.add(ParagraphStyle(
            'RightAligned',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_RIGHT,
            fontName='Helvetica'
        ))
        
        # Center aligned
        styles.add(ParagraphStyle(
            'CenterAligned',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            fontName='Helvetica'
        ))
        
        return styles


class TranscriptPDFGenerator:
    """Generate official academic transcripts"""
    
    def __init__(self, student):
        self.student = student
        self.styles = PDFGenerator.create_styles()
        
    def generate(self, buffer=None):
        """Generate PDF transcript"""
        if buffer is None:
            buffer = io.BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        elements = []
        
        # Header
        elements.extend(self._create_header())
        elements.append(Spacer(1, 20))
        
        # Student Information
        elements.extend(self._create_student_info())
        elements.append(Spacer(1, 20))
        
        # Academic Summary
        elements.extend(self._create_academic_summary())
        elements.append(Spacer(1, 20))
        
        # Course History by Semester
        elements.extend(self._create_course_history())
        
        # Footer
        elements.extend(self._create_footer())
        
        # Build PDF
        doc.build(elements)
        
        if buffer is None:
            return None
        
        buffer.seek(0)
        return buffer
    
    def _create_header(self):
        """Create transcript header"""
        elements = []
        
        # University Name
        elements.append(Paragraph(
            "UNIVERSITY SYSTEM",
            self.styles['CustomTitle']
        ))
        
        elements.append(Paragraph(
            "Official Academic Transcript",
            self.styles['CustomHeading2']
        ))
        
        elements.append(Spacer(1, 10))
        
        # Transcript Info
        info_data = [
            [Paragraph(_("Transcript ID: TR-"), self.styles['Small']),
             Paragraph(f"{self.student.student_no}-{datetime.now().strftime('%Y%m%d')}", self.styles['Small'])],
            [Paragraph(_("Issue Date: "), self.styles['Small']),
             Paragraph(datetime.now().strftime('%B %d, %Y'), self.styles['Small'])],
        ]
        
        info_table = Table(info_data, colWidths=[4*cm, 6*cm])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        
        elements.append(info_table)
        
        return elements
    
    def _create_student_info(self):
        """Create student information section"""
        elements = []
        
        elements.append(Paragraph(
            "Student Information",
            self.styles['CustomHeading2']
        ))
        
        # Student details table
        data = [
            [Paragraph(_("Student Name:"), self.styles['CustomNormal']),
             Paragraph(self.student.get_full_name(), self.styles['CustomNormal'])],
            [Paragraph(_("Student ID:"), self.styles['CustomNormal']),
             Paragraph(self.student.student_no, self.styles['CustomNormal'])],
            [Paragraph(_("Program:"), self.styles['CustomNormal']),
             Paragraph(str(self.student.program.name_en) if self.student.program else '-', self.styles['CustomNormal'])],
            [Paragraph(_("Department:"), self.styles['CustomNormal']),
             Paragraph(str(self.student.department.name_en) if self.student.department else '-', self.styles['CustomNormal'])],
            [Paragraph(_("Admission Date:"), self.styles['CustomNormal']),
             Paragraph(self.student.admission_date.strftime('%B %d, %Y') if self.student.admission_date else '-', self.styles['CustomNormal'])],
        ]
        
        table = Table(data, colWidths=[4*cm, 10*cm])
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        elements.append(table)
        
        return elements
    
    def _create_academic_summary(self):
        """Create academic summary section"""
        elements = []
        
        elements.append(Paragraph(
            "Academic Summary",
            self.styles['CustomHeading2']
        ))
        
        # Calculate totals with repeated course handling
        enrollments = Enrollment.objects.filter(
            student=self.student,
            final_result__is_published=True
        ).select_related('final_result', 'course_offering__course')
        
        # Track best grades for repeated courses
        course_best_grades = {}
        for enrollment in enrollments:
            course_id = enrollment.course_offering.course_id
            grade_point = enrollment.final_result.grade_point or Decimal('0')
            credits = Decimal(str(enrollment.course_offering.course.credit_hours))
            
            if course_id not in course_best_grades or grade_point > course_best_grades[course_id]['grade_point']:
                course_best_grades[course_id] = {
                    'grade_point': grade_point,
                    'credits': credits
                }
        
        total_credits = sum(d['credits'] for d in course_best_grades.values())
        total_quality_points = sum(d['grade_point'] * d['credits'] for d in course_best_grades.values())
        
        if total_credits > 0:
            cgpa = (total_quality_points / total_credits).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            cgpa = Decimal('0.00')
        
        # Summary data
        data = [
            [Paragraph(_("Cumulative GPA (CGPA):"), self.styles['CustomNormal']),
             Paragraph(str(cgpa), self.styles['CustomNormal'])],
            [Paragraph(_("Total Credits Earned:"), self.styles['CustomNormal']),
             Paragraph(str(total_credits), self.styles['CustomNormal'])],
            [Paragraph(_("Academic Standing:"), self.styles['CustomNormal']),
             Paragraph(_("Good Standing") if cgpa >= 2.0 else _("Academic Probation"), self.styles['CustomNormal'])],
        ]
        
        table = Table(data, colWidths=[4*cm, 10*cm])
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        elements.append(table)
        
        return elements
    
    def _create_course_history(self):
        """Create course history by semester"""
        elements = []
        
        elements.append(Paragraph(
            "Course History",
            self.styles['CustomHeading2']
        ))
        
        # Get completed enrollments grouped by semester
        enrollments = Enrollment.objects.filter(
            student=self.student,
            final_result__is_published=True
        ).select_related(
            'course_offering',
            'course_offering__course',
            'course_offering__semester',
            'final_result'
        ).order_by('course_offering__semester__start_date')
        
        if not enrollments:
            elements.append(Paragraph(
                "No completed courses on record.",
                self.styles['CustomNormal']
            ))
            return elements
        
        # Group by semester
        from collections import defaultdict
        semester_courses = defaultdict(list)
        for e in enrollments:
            semester_name = e.course_offering.semester.display_name if e.course_offering.semester else _('Unknown')
            semester_courses[semester_name].append(e)
        
        # Create table for each semester
        for semester, courses in semester_courses.items():
            elements.append(Spacer(1, 10))
            elements.append(Paragraph(
                f"{semester}",
                self.styles['CustomNormal']
            ))
            
            # Course table
            course_data = [
                [_('Course Code'), _('Course Title'), _('Credits'), _('Grade'), _('Grade Point'), _('Status')]
            ]
            
            for course in courses:
                course_data.append([
                    course.course_offering.course.code,
                    course.course_offering.course.get_title(),
                    str(course.course_offering.course.credit_hours),
                    course.final_result.letter_grade if course.final_result else '-',
                    str(course.final_result.grade_point) if course.final_result else '-',
                    _('Pass') if course.pass_fail_status == 'pass' else _('Fail')
                ])
            
            course_table = Table(course_data, colWidths=[2.5*cm, 6*cm, 1.5*cm, 2*cm, 2*cm, 2*cm])
            course_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e5e7eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
            ]))
            
            elements.append(course_table)
        
        return elements
    
    def _create_footer(self):
        """Create transcript footer"""
        elements = []
        
        elements.append(Spacer(1, 30))
        
        # Disclaimer
        elements.append(Paragraph(
            "This is an official transcript issued by the University System.",
            self.styles['Small']
        ))
        
        elements.append(Paragraph(
            "Any alteration or reproduction invalidates this document.",
            self.styles['Small']
        ))
        
        elements.append(Spacer(1, 20))
        
        # Signature line
        elements.append(Paragraph(
            "_______________________________",
            self.styles['CenterAligned']
        ))
        elements.append(Paragraph(
            "Registrar",
            self.styles['CenterAligned']
        ))
        
        return elements


class ReportGenerator:
    """Generate various academic reports"""
    
    @staticmethod
    def generate_semester_report(semester, format='excel'):
        """Generate semester results report"""
        enrollments = Enrollment.objects.filter(
            course_offering__semester=semester,
            final_result__is_published=True
        ).select_related(
            'student',
            'course_offering',
            'course_offering__course',
            'final_result'
        )
        
        if format == 'excel':
            return ReportGenerator._generate_excel_report(enrollments, semester)
        elif format == 'csv':
            return ReportGenerator._generate_csv_report(enrollments, semester)
        else:
            raise ValueError(f"Unsupported format: {format}")
    
    @staticmethod
    def _generate_excel_report(enrollments, semester):
        """Generate Excel report"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Semester Results"
        
        # Header
        ws['A1'] = f"Semester Results Report - {semester.display_name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Column headers
        headers = ['Student ID', 'Student Name', 'Course Code', 'Course Title', 
                   'Credits', 'Grade', 'Grade Point', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        row = 4
        for enrollment in enrollments:
            ws.cell(row=row, column=1, value=enrollment.student.student_no)
            ws.cell(row=row, column=2, value=enrollment.student.get_full_name())
            ws.cell(row=row, column=3, value=enrollment.course_offering.course.code)
            ws.cell(row=row, column=4, value=enrollment.course_offering.course.get_title())
            ws.cell(row=row, column=5, value=enrollment.course_offering.course.credit_hours)
            ws.cell(row=row, column=6, value=enrollment.final_result.letter_grade if enrollment.final_result else '-')
            ws.cell(row=row, column=7, value=float(enrollment.final_result.grade_point) if enrollment.final_result else 0)
            ws.cell(row=row, column=8, value='Pass' if enrollment.pass_fail_status == 'pass' else 'Fail')
            row += 1
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="semester_results_{semester.id}.xlsx"'
        
        wb.save(response)
        return response
    
    @staticmethod
    def _generate_csv_report(enrollments, semester):
        """Generate CSV report"""
        import csv
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="semester_results_{semester.id}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student ID', 'Student Name', 'Course Code', 'Course Title', 
                        'Credits', 'Grade', 'Grade Point', 'Status'])
        
        for enrollment in enrollments:
            writer.writerow([
                enrollment.student.student_no,
                enrollment.student.get_full_name(),
                enrollment.course_offering.course.code,
                enrollment.course_offering.course.get_title(),
                enrollment.course_offering.course.credit_hours,
                enrollment.final_result.letter_grade if enrollment.final_result else '-',
                float(enrollment.final_result.grade_point) if enrollment.final_result else 0,
                'Pass' if enrollment.pass_fail_status == 'pass' else 'Fail'
            ])
        
    @staticmethod
    def generate_transcript_excel(student):
        """Generate student transcript as Excel"""
        from collections import defaultdict
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Transcript"
        
        # Header
        ws['A1'] = "UNIVERSITY SYSTEM - OFFICIAL TRANSCRIPT"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Student Info
        ws['A3'] = "Student Name:"
        ws['B3'] = student.get_full_name()
        ws['A4'] = "Student ID:"
        ws['B4'] = student.student_no
        ws['A5'] = "Program:"
        ws['B5'] = str(student.program.name_en) if student.program else '-'
        ws['A6'] = "Issue Date:"
        ws['B6'] = datetime.now().strftime('%B %d, %Y')
        
        # Get enrollments grouped by semester with repeated course handling
        enrollments = Enrollment.objects.filter(
            student=student,
            final_result__is_published=True
        ).select_related(
            'course_offering__course',
            'course_offering__semester',
            'final_result'
        ).order_by('course_offering__semester__start_date')
        
        semester_data = defaultdict(lambda: {'courses': [], 'total_credits': 0, 'total_points': Decimal('0')})
        course_best_grades = {}
        
        for enrollment in enrollments:
            semester = enrollment.course_offering.semester
            course = enrollment.course_offering.course
            result = enrollment.final_result
            
            if not result:
                continue
            
            grade_points = result.grade_point or Decimal('0')
            credits = course.credit_hours
            
            course_key = (semester.id, course.id)
            if course_key not in course_best_grades or grade_points > course_best_grades[course_key]['grade_points']:
                course_best_grades[course_key] = {
                    'enrollment': enrollment,
                    'grade_points': grade_points,
                    'credits': credits
                }
        
        for course_key, data in course_best_grades.items():
            semester = data['enrollment'].course_offering.semester
            semester_data[semester]['courses'].append(data['enrollment'])
            semester_data[semester]['total_credits'] += data['credits']
            semester_data[semester]['total_points'] += data['grade_points'] * data['credits']
        
        # Start row for course data
        current_row = 8
        
        cumulative_points = Decimal('0')
        cumulative_credits = Decimal('0')
        
        for semester in sorted(semester_data.keys(), key=lambda s: s.start_date):
            data = semester_data[semester]
            
            if data['total_credits'] > 0:
                gpa = (data['total_points'] / data['total_credits']).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            else:
                gpa = Decimal('0.00')
            
            cumulative_points += data['total_points']
            cumulative_credits += data['total_credits']
            
            if cumulative_credits > 0:
                cgpa = (cumulative_points / cumulative_credits).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            else:
                cgpa = Decimal('0.00')
            
            # Semester header
            ws.merge_cells(f'A{current_row}:F{current_row}')
            cell = ws.cell(row=current_row, column=1, value=f"{semester.display_name} - GPA: {gpa} | CGPA: {cgpa}")
            cell.font = Font(bold=True, size=12, color='FFFFFF')
            cell.fill = PatternFill(start_color='1e40af', end_color='1e40af', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            current_row += 1
            
            # Column headers
            headers = ['Course Code', 'Course Title', 'Credits', 'Grade', 'Points', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            current_row += 1
            
            # Course rows
            for enrollment in data['courses']:
                ws.cell(row=current_row, column=1, value=enrollment.course_offering.course.code)
                ws.cell(row=current_row, column=2, value=enrollment.course_offering.course.get_title())
                ws.cell(row=current_row, column=3, value=enrollment.course_offering.course.credit_hours)
                ws.cell(row=current_row, column=4, value=enrollment.final_result.letter_grade)
                ws.cell(row=current_row, column=5, value=float(enrollment.final_result.grade_point) if enrollment.final_result.grade_point is not None else 0.0)
                ws.cell(row=current_row, column=6, value='Pass' if enrollment.pass_fail_status == 'pass' else 'Fail')
                current_row += 1
            
            current_row += 1
        
        # Final summary
        ws.merge_cells(f'A{current_row}:F{current_row}')
        cell = ws.cell(row=current_row, column=1, value=f"FINAL CGPA: {cgpa} | Total Credits: {cumulative_credits}")
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="transcript_{student.student_no}.xlsx"'
        
        wb.save(response)
        return response

    @staticmethod
    def get_grade_distribution(cohort=None, program=None, semester=None):
        """Get grade distribution for analysis"""
