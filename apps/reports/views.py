"""
Report Generation Views
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.http import HttpResponse
from django.contrib import messages
from django.utils.translation import gettext_lazy as _

from apps.core.services import admin_required, teacher_required
from apps.core.report_services import (
    TranscriptPDFGenerator, ReportGenerator
)
from apps.core.academic_services import GradeCalculationService
from apps.students.models import Student
from apps.semesters.models import Semester
from apps.courses.offering_models import CourseOffering
from apps.enrollments.models import Enrollment
from apps.warnings.models import EarlyWarningResult
import csv


class GenerateTranscriptView(View):
    """Generate student transcript PDF"""
    
    @method_decorator(login_required)
    @method_decorator(admin_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request, student_id):
        student = get_object_or_404(Student, pk=student_id)
        
        # Generate PDF
        generator = TranscriptPDFGenerator(student)
        pdf_buffer = generator.generate()
        
        if pdf_buffer is None:
            messages.error(request, _('Error generating transcript'))
            return redirect('admin_portal:admin_students')
        
        # Create response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="transcript_{student.student_no}.pdf"'
        response.write(pdf_buffer.getvalue())
        pdf_buffer.close()
        
        return response


class GenerateSemesterReportView(View):
    """Generate semester performance report"""
    
    @method_decorator(login_required)
    @method_decorator(admin_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request, student_id, semester_id):
        semester = get_object_or_404(Semester, pk=semester_id)
        format = request.GET.get('format', 'excel')
        
        if format not in ['excel', 'csv']:
            format = 'excel'
        
        return ReportGenerator.generate_semester_report(semester, format)


class GenerateCourseReportView(View):
    """Generate course performance report"""
    
    @method_decorator(login_required)
    @method_decorator(admin_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request, offering_id):
        offering = get_object_or_404(CourseOffering, pk=offering_id)
        
        # Get all enrollments for this offering
        enrollments = Enrollment.objects.filter(
            course_offering=offering,
            final_result__is_published=True
        ).select_related('student', 'final_result')
        
        if not enrollments:
            messages.warning(request, _('No published results found for this course'))
            return redirect('admin_portal:admin_offerings')
        
        # Generate Excel report
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Course Report"
        
        # Title
        ws['A1'] = f"Course Performance Report - {offering.course.code}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Semester: {offering.semester.display_name if offering.semester else 'N/A'}"
        ws.merge_cells('A2:F2')
        
        # Headers
        headers = ['Student ID', 'Student Name', 'Grade', 'Grade Point', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data
        row = 5
        for enrollment in enrollments:
            ws.cell(row=row, column=1, value=enrollment.student.student_no)
            ws.cell(row=row, column=2, value=enrollment.student.get_full_name())
            ws.cell(row=row, column=3, value=enrollment.final_result.letter_grade if enrollment.final_result else '-')
            ws.cell(row=row, column=4, value=float(enrollment.final_result.grade_point) if enrollment.final_result else 0)
            ws.cell(row=row, column=5, value='Pass' if enrollment.pass_fail_status == 'pass' else 'Fail')
            row += 1
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="course_report_{offering.course.code}.xlsx"'
        
        wb.save(response)
        return response


class GenerateWarningReportView(View):
    """Generate warning summary report"""
    
    @method_decorator(login_required)
    @method_decorator(admin_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request):
        # Get all warning results
        warnings = EarlyWarningResult.objects.select_related(
            'student',
            'semester'
        ).order_by('-generated_at')
        
        # Generate Excel report
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Warning Summary"
        
        # Title
        ws['A1'] = "Academic Warning Summary Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:G1')
        
        # Headers
        headers = ['Student ID', 'Student Name', 'Semester', 'Warning Level', 
                   'Risk Score', 'Generated Date', 'Acknowledged']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data
        row = 4
        for warning in warnings:
            ws.cell(row=row, column=1, value=warning.student.student_no)
            ws.cell(row=row, column=2, value=warning.student.get_full_name())
            ws.cell(row=row, column=3, value=warning.semester.display_name if warning.semester else 'N/A')
            ws.cell(row=row, column=4, value=warning.warning_level)
            ws.cell(row=row, column=5, value=float(warning.risk_score) if warning.risk_score else 0)
            ws.cell(row=row, column=6, value=warning.generated_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=7, value='Yes' if warning.is_acknowledged else 'No')
            row += 1
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="warning_summary_report.xlsx"'
        
        wb.save(response)
        return response


class GenerateAttendanceReportView(View):
    """Generate attendance report"""
    
    @method_decorator(login_required)
    @method_decorator(teacher_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request, offering_id):
        offering = get_object_or_404(CourseOffering, pk=offering_id)
        
        # Get all enrollments for this offering
        enrollments = Enrollment.objects.filter(
            course_offering=offering
        ).select_related('student')
        
        # Generate CSV report
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="attendance_report_{offering.course.code}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student ID', 'Student Name', 'Total Classes', 
                        'Attended Classes', 'Attendance %', 'Status'])
        
        for enrollment in enrollments:
            attendance_pct = enrollment.attendance_percentage or 0
            status = 'Good' if attendance_pct >= 80 else 'Warning' if attendance_pct >= 60 else 'Critical'
            
            writer.writerow([
                enrollment.student.student_no,
                enrollment.student.get_full_name(),
                enrollment.total_classes,
                enrollment.attended_classes,
                attendance_pct,
                status
            ])
        
        return response


class ExportStudentsView(View):
    """Export students to Excel/CSV"""
    
    @method_decorator(login_required)
    @method_decorator(admin_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        students = Student.objects.select_related('user', 'department', 'program').all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Title
        ws['A1'] = "Student Export"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:I1')
        
        # Headers
        headers = ['Student ID', 'Full Name', 'Email', 'Department', 'Program', 
                   'Batch Year', 'CGPA', 'Status', 'Enrollment Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data
        row = 4
        for student in students:
            ws.cell(row=row, column=1, value=student.student_no)
            ws.cell(row=row, column=2, value=student.get_full_name())
            ws.cell(row=row, column=3, value=student.user.email if student.user else '')
            ws.cell(row=row, column=4, value=str(student.department) if student.department else '')
            ws.cell(row=row, column=5, value=str(student.program) if student.program else '')
            ws.cell(row=row, column=6, value=student.batch_year)
            ws.cell(row=row, column=7, value=float(student.cgpa) if student.cgpa else 0)
            ws.cell(row=row, column=8, value=student.status)
            ws.cell(row=row, column=9, value=student.enrollment_date.strftime('%Y-%m-%d') if student.enrollment_date else '')
            row += 1
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="students_export.xlsx"'
        
        wb.save(response)
        return response


class ExportGradesView(View):
    """Export grades to Excel/CSV"""
    
    @method_decorator(login_required)
    @method_decorator(admin_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        enrollments = Enrollment.objects.filter(
            final_result__is_published=True
        ).select_related('student', 'course_offering', 'course_offering__course', 'final_result')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Grades"
        
        # Title
        ws['A1'] = "Grades Export"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = ['Student ID', 'Student Name', 'Course Code', 'Course Title',
                   'Grade', 'Grade Point', 'Credits', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data
        row = 4
        for enrollment in enrollments:
            ws.cell(row=row, column=1, value=enrollment.student.student_no)
            ws.cell(row=row, column=2, value=enrollment.student.get_full_name())
            ws.cell(row=row, column=3, value=enrollment.course_offering.course.code)
            ws.cell(row=row, column=4, value=enrollment.course_offering.course.get_title())
            ws.cell(row=row, column=5, value=enrollment.final_result.letter_grade if enrollment.final_result else '-')
            ws.cell(row=row, column=6, value=float(enrollment.final_result.grade_point) if enrollment.final_result else 0)
            ws.cell(row=row, column=7, value=enrollment.course_offering.course.credit_hours)
            ws.cell(row=row, column=8, value='Pass' if enrollment.pass_fail_status == 'pass' else 'Fail')
            row += 1
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="grades_export.xlsx"'
        
        wb.save(response)
        return response


class ExportEnrollmentsView(View):
    """Export enrollments to Excel/CSV"""
    
    @method_decorator(login_required)
    @method_decorator(admin_required)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        enrollments = Enrollment.objects.select_related(
            'student', 'course_offering', 'course_offering__course', 'course_offering__semester'
        ).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Enrollments"
        
        # Title
        ws['A1'] = "Enrollments Export"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:H1')
        
        # Headers
        headers = ['Student ID', 'Student Name', 'Course Code', 'Course Title',
                   'Semester', 'Status', 'Enrolled Date', 'Final Grade']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data
        row = 4
        for enrollment in enrollments:
            ws.cell(row=row, column=1, value=enrollment.student.student_no)
            ws.cell(row=row, column=2, value=enrollment.student.get_full_name())
            ws.cell(row=row, column=3, value=enrollment.course_offering.course.code)
            ws.cell(row=row, column=4, value=enrollment.course_offering.course.get_title())
            ws.cell(row=row, column=5, value=enrollment.course_offering.semester.display_name if enrollment.course_offering.semester else 'N/A')
            ws.cell(row=row, column=6, value=enrollment.get_enroll_status_display())
            ws.cell(row=row, column=7, value=enrollment.enrolled_at.strftime('%Y-%m-%d') if enrollment.enrolled_at else '')
            ws.cell(row=row, column=8, value=enrollment.final_result.letter_grade if enrollment.final_result else '-')
            row += 1
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="enrollments_export.xlsx"'
        
        wb.save(response)
        return response
